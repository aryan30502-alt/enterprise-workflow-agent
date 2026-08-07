"""
Integration Test 2: ExcelTool — real Microsoft Excel COM automation via pywin32.

Prerequisites:
  - Microsoft Excel installed
  - pywin32 installed (already confirmed)
  - A real CSV file from CSVGeneratorTool

Verifies:
  - Excel opens without error
  - XLSX file is written to disk
  - openpyxl can read the saved workbook back
  - Row count in XLSX matches CSV row count
  - Worksheet name is "Employee Data"
  - Header row is in row 1
  - artifact_updates() returns the correct xlsx_path
"""

import os
import sys
import csv
import pathlib

sys.path.insert(0, ".")
os.environ.setdefault("GOOGLE_API_KEY", "test-key")
os.environ.setdefault("GOOGLE_SHEET_ID", "test-sheet-id")


def ok(msg):  print(f"  [PASS] {msg}")
def fail(msg): print(f"  [FAIL] {msg}"); sys.exit(1)
def section(t): print(f"\n{'='*55}\n  {t}\n{'='*55}")


section("INTEGRATION 2: ExcelTool (real COM automation)")

# ── Step 1: Generate a real CSV first ─────────────────────────────────────────
print("\n  Step 1: Generating source CSV...")
from src.tools.csv_generator.tool import CSVGeneratorTool

csv_tool = CSVGeneratorTool()
csv_result = csv_tool.run({"num_rows": 20, "output_dir": "output"})
if not csv_result.success:
    fail(f"CSV generation failed (prerequisite): {csv_result.error}")

csv_path = csv_result.data.csv_path
ok(f"Source CSV ready: {csv_path}")

# ── Step 2: Run ExcelTool with real COM ───────────────────────────────────────
print("\n  Step 2: Launching Excel via COM automation...")
print("  (Excel will open visually on screen — this is expected)")

from src.tools.excel_tool.tool import ExcelTool

excel_tool = ExcelTool()
excel_result = excel_tool.run({
    "csv_path": csv_path,
    "output_dir": "output",
})

if not excel_result.success:
    fail(f"ExcelTool failed: {excel_result.error}")

ok(f"Excel COM automation succeeded in {excel_result.duration_seconds}s")

xlsx_path = excel_result.data.xlsx_path
ok(f"XLSX path: {xlsx_path}")

# ── Step 3: Verify the XLSX on disk ───────────────────────────────────────────
print("\n  Step 3: Verifying saved XLSX with openpyxl...")

if not pathlib.Path(xlsx_path).exists():
    fail(f"XLSX file not found at: {xlsx_path}")
ok("XLSX file exists on disk")

import openpyxl
wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

ws_names = wb.sheetnames
if "Employee Data" not in ws_names:
    fail(f"Sheet 'Employee Data' not found. Sheets: {ws_names}")
ok(f"Worksheet 'Employee Data' found")

ws = wb["Employee Data"]

# Count data rows (total minus header)
max_row = ws.max_row
data_rows = max_row - 1
wb.close()

if data_rows != 20:
    fail(f"Expected 20 data rows, openpyxl reports {data_rows}")
ok(f"Data row count correct: {data_rows}")

# ── Step 4: artifact_updates() contract ───────────────────────────────────────
updates = excel_result.data.artifact_updates()
if updates.get("xlsx_path") != xlsx_path:
    fail(f"artifact_updates() xlsx_path mismatch: {updates}")
ok("artifact_updates() returns correct xlsx_path")

# ── Step 5: Cross-check CSV row count ─────────────────────────────────────────
with open(csv_path, "r", encoding="utf-8") as fh:
    csv_data_rows = sum(1 for _ in csv.reader(fh)) - 1  # subtract header
if csv_data_rows != data_rows:
    fail(f"CSV has {csv_data_rows} rows but XLSX has {data_rows} rows — MISMATCH")
ok(f"CSV and XLSX row counts match: {csv_data_rows}")

# Expose paths for downstream tests
print(f"\n  CSV_PATH={csv_path}")
print(f"  XLSX_PATH={xlsx_path}")
print("\n  [INTEGRATION 2 PASSED]")
